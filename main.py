import datetime
from datetime import time
#importa el instante del tiempo (hora)
import openpyxl
from openpyxl import Workbook
g="./documento1.xlsx"
e=Workbook()
lk=e.active
lk["A1"]=datetime.datetime.now().time()
lk["B1"]=datetime.time(4,23,4,5)
lk["C1"]=datetime.time(5,2,6,7)
lk["D1"]=datetime.time(3,2,4,9)
w=e.save(g)
#crear el documento, el cual utilizare para importar,en este caso se llama (documento1.xlsx). 

from openpyxl import load_workbook
hy=load_workbook("documento1.xlsx")
#importa el documento1
t=hy.active
e=lk["E1"]="=SUMA(B1:C1)"
f=lk["F1"]="=SUMA(A1;B1;C1;D1)"
#añado celdas en formato de horas a el documento1. 
hy.save("resultado1 documento1")
#guardo todas las operaciones y modificaciones del documento1 en otro documento(resultado documento1.)

from openpyxl import load_workbook
q=load_workbook("documento1.xlsx")
rt=q.active
#importa el documento1
rt["A2"]=datetime.time(0,2,5,67)
rt["B2"]=datetime.time(3,4,5,7)
rt["C2"]= datetime.datetime.utcnow().time()
q.save("hora.xlsx")
#escribo otras celdas en formato de hora y hago operaciones. Tambien creo el nuevo documento para guardar todo el proceso. 
y=q.active
A1=y["A1"].value
B1=y["B1"].value
B2=y["B2"].value
F1=y["F1"].value
E1=y["E1"].value
C1=y["C1"].value
C2=y["C2"].value
celdas=[A1,B1,B2,F1,E1,C1,C2]
for valor in celdas:
  print(valor)
#muestra algunos valores de las celdas, E1 Y F1 no aparecen por que estan guardadas en el documento=resultado1. 